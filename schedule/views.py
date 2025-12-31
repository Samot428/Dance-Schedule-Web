from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.views.decorators.csrf import ensure_csrf_cookie
from django.urls import reverse
from .models import UploadedScheduleFile
from main.models import DancersAvailability, Day, Dancer, Group, Couple, Trainer, TrainerDayAvailability
from .reading_excel_func import read_dancers_availability
from openpyxl import load_workbook
import os
from collections import defaultdict
from .making_schedule_func import (
    convert_to_min,
    min_to_time_str,
    make_trainers_windows,
    sort_couples_by_group,
    create_schedule as build_schedule,
)
# Create your views here.
def calendar_view(request):
    return redirect('main:calendar_view')

@ensure_csrf_cookie
def schedule_view(request):
    return render(request, 'schedule_view.html')

def validate_excel_format(file_path):
    """Validate if the Excel file has the correct format for reading_excel_func.py"""
    try:
        # Check if file exists and is a valid Excel file
        wb = load_workbook(file_path)

        # Check if at least one group sheet exists
        groups = Group.objects.all()
        if not groups.exists():
            return False, "No groups defined in the database. Please define groups first." 

        # Validate that at least one group sheet exists in the workbook
        found_valid_sheet = False
        for group in groups:
            if group.name in wb.sheetnames:
                found_valid_sheet = True
                try:
                    # Try to read availability from this sheet
                    day_times, day_dancers_avail = read_dancers_availability(group.name, file_path)

                    # Check if data was extracted
                    if not day_times or not day_dancers_avail:
                        return False, f'Sheet "{group.name}" exists but has no valid data structure.'

                except Exception as e:
                    return False, f'Sheet "{group.name}" format error: {str(e)}'
        if not found_valid_sheet:
            available_sheets = ", ".join(wb.sheetnames)
            expected_sheets = ", ".join([g.name for g in groups])
            return False, f"No valid group sheets found. Expected: {expected_sheets}, Found: {available_sheets}" 
        return True, 'File format is valid'
    except Exception as e:
        return False, f'Error reading file: {str(e)}'

def upload_schedule_files(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=400)

    files = request.FILES.getlist('files') or []
    if not files:
        return JsonResponse({'status': 'error', 'message': 'No files provided'}, status=400)
        
    uploaded_files_info = []
    for f in files:
        uploaded_file = UploadedScheduleFile.objects.create(
            filename=f.name,
            file=f,
        )

        # Validate the uploaded file format
        is_valid, error_message = validate_excel_format(uploaded_file.file.path)

        if not is_valid:
            # Delete invalid file and return error
            uploaded_file.file.delete(save=False)
            uploaded_file.delete()
            return JsonResponse({
                'status':'error',
                'message':f'File "{f.name}" has invalid format: {error_message}'
            }, status=400)

        uploaded_files_info.append({
            'id': uploaded_file.id,
            'filename': uploaded_file.filename,
            'size': uploaded_file.file.size,
            'uploaded_at': uploaded_file.uploaded_at.strftime('%Y-%m-%d %H:%M'),
            'url': uploaded_file.file.url,  # useful for preview/download
        })

    return JsonResponse({'status': 'success', 'files': uploaded_files_info})

def get_uploaded_files(request):
    files = UploadedScheduleFile.objects.all().order_by('-uploaded_at')
    files_data = [{
        'id': f.id,
        'filename': f.filename,
        'size': f.file.size,
        'uploaded_at': f.uploaded_at.strftime('%Y-%m-%d %H:%M'),
        'url': f.file.url,
    } for f in files]
    return JsonResponse({'files': files_data})

from django.views.decorators.csrf import csrf_exempt

def delete_uploaded_file(request, file_id):
    if request.method not in ('DELETE', 'POST'):
        return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=400)

    try:
        f = UploadedScheduleFile.objects.get(id=file_id)
    except UploadedScheduleFile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'File not found'}, status=404)

    f.file.delete(save=False)
    f.delete()
    return JsonResponse({'status': 'success'})

def load_dancers_availability(file_id):
    """Load dancer's availability from the uploaded Excel file"""
    # if request.method != 'POST':
    #     return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=400)
    uploaded_file = UploadedScheduleFile.objects.filter(id=file_id).first()
    
    for group in Group.objects.all():
        day_times, day_dancers_avail = read_dancers_availability(group.name, uploaded_file.file.path)
        for day_name, dancers_availability in day_dancers_avail.items():
            day_obj, _ = Day.objects.get_or_create(name=day_name)
            for dancer_name, availability in dancers_availability.items():
                dancer_obj, _ = Dancer.objects.get_or_create(name=dancer_name)
                DancersAvailability.objects.update_or_create(
                    dancer=dancer_obj,
                    day=day_obj,
                    defaults={'availability': availability}
                )

def create_schedule(request):
    """Create schedule for all the days that were inserted in the database by user"""
    if request.method == 'POST':
        file_id = request.POST.get('file_id')
        sort_couples_by = request.POST.get('sort_by', 'Group Index')
        uploaded_file = UploadedScheduleFile.objects.get(id=file_id)

        dawt = defaultdict(list)
        cawt = defaultdict(list)
        for group in Group.objects.all():
            day_times, dancers_avail = read_dancers_availability(group.name, uploaded_file.file.path)

            # compute dancers availablity with times in format (time_str, True/False)
            
            for day in dancers_avail:
                dancers = {}
                for d in dancers_avail[day]:
                    avail = []
                    for i in range(len(day_times[day])):
                        time = day_times[day][i]
                        time = convert_to_min(time)
                        time_str = min_to_time_str(time)
                        avail.append((time_str, dancers_avail[day][d][i]))
                    dancers[d] = avail
                dawt[day].append(dancers)
            
            # couples availability with times
            couples = group.couples.all()
            all_schedules = {}
            for day in day_times:
                day_couples = {}
                dancers_for_day = dawt[day][-1] if dawt[day] else {}
                for couple in couples:
                    name = couple.name
                    dancer1 = name.split(' a ')[0] if ' a ' in name else name.split(' and ')[0]
                    dancer2 = name.split(' a ')[1] if ' a ' in name else name.split(' and ')[1]
                    avail = []
                    for i in range(len(day_times[day])):
                        time = day_times[day][i]
                        time = convert_to_min(time)
                        time_str = min_to_time_str(time)
                        d1_slots = dancers_for_day.get(dancer1, [])
                        d2_slots = dancers_for_day.get(dancer2, [])
                        has_d1 = i < len(d1_slots)
                        has_d2 = i < len(d2_slots)
                        available = bool(has_d1 and has_d2 and d1_slots[i][1] and d2_slots[i][1])
                        avail.append((time_str, available))
                    day_couples[name] = avail
                cawt[day].append(day_couples)
        all_schedules = {}
        for day in Day.objects.all():
            trainers = day.trainers.all()

            if not trainers.exists():
                # add maybe something like a jsonresponse
                print(f'No trainers')
                continue

            couples = day.couples.all()

            if not couples.exists():
                # add maybe something like a jsonresponse
                print(f'No couples')
                continue

            try:
                
                tw = make_trainers_windows(trainers, day)
                for trainer, windows in tw.items():
                    available_slots = sum(1 for _, is_avail in windows if is_avail)
            except Exception as e:
                print(f'Error creating trainer windows: {e}')
                continue

            if sort_couples_by == 'group':
                sorted_couples = sort_couples_by_group(couples, Group.objects.all())
            else:
                sorted_couples = list(couples)
            schedule = build_schedule(cawt=cawt[day.name][0], trainers_windows=tw, couples=sorted_couples)
            if schedule:
                all_schedules[day.name] = schedule
        if not all_schedules:
            return JsonResponse({
                 'status':'error',
                 'message':'Could not create schedule for any configured day'
            })

        formatted_schedule = {}
        for day_name, day_schedule in all_schedules.items():
            lessons = []
            for couple, (trainer, time_str) in day_schedule.items():
                lessons.append({
                    'couple':couple.name,
                    'trainer':trainer.name,
                    'time':time_str,
                    'duration':couple.min_duration,
                })
            # Sort lessons by time chronologically
            lessons.sort(key=lambda x: tuple(map(int, x['time'].split(':'))))
            formatted_schedule[day_name] = lessons
        
        return JsonResponse({
            'status':'success',
            'message':'Schedule created successfully',
            'schedule': formatted_schedule
        })
    return JsonResponse({
        'status':'error',
        'message':'Invalid request method'
    }, status=400)

# def filter_times_by_day_hours(day_times_list, day_availability_lists, day_start_time, day_end_time):
#     """
#     Filter times and availability to only include times within the day's start and end times.
    
#     Args:
#         day_times_list: list of time objects
#         day_availability_lists: list of availability lists (one per time slot)
#         day_start_time: datetime.time - day's start time
#         day_end_time: datetime.time - day's end time
    
#     Returns:
#         (filtered_times, filtered_availabilities)
#     """
#     filtered_times = []
#     filtered_availabilities = []
    
#     for i, t in enumerate(day_times_list):
#         if day_start_time <= t <= day_end_time:
#             filtered_times.append(t)
#             filtered_availabilities.append(day_availability_lists[i])
    
#     return filtered_times, filtered_availabilities

# def create_schedule(request):
#     """Creates the schedule for user-defined days only"""
#     if request.method == 'POST':
#         file_id = request.POST.get('file_id')
#         sort_couple_by = request.POST.get('sort_by', 'group')

#         if not file_id:
#             return JsonResponse({
#                 'status': 'error',
#                 'message': 'No file ID provided'
#             }, status=400)

#         try:
#             uploaded_file = UploadedScheduleFile.objects.get(id=file_id)
#         except UploadedScheduleFile.DoesNotExist:
#             return JsonResponse({
#                 'status': 'error',
#                 'message': 'File not found'
#             }, status=400)

#         all_schedules = {}

#         # Only process days that exist in the database
#         user_days = Day.objects.all()
        
#         if not user_days.exists():
#             return JsonResponse({
#                 'status': 'error',
#                 'message': 'No days configured in the system. Please create days first.'
#             }, status=400)

#         print(f"\nProcessing {user_days.count()} configured days")
#         print(f"Days: {[d.name for d in user_days]}\n")

#         for day in user_days:
#             trainers = day.trainers.all()

#             if not trainers.exists():
#                 print(f"Skipping {day.name}: no trainers assigned")
#                 continue

#             couples = day.couples.all()

#             if not couples.exists():
#                 print(f"Skipping {day.name}: no couples assigned")
#                 continue

#             print(f"\n{'='*60}")
#             print(f"Processing Day: {day.name}")
#             print(f"Day Hours: {day.start_time} - {day.end_time}")
#             print(f"Couples: {[c.name for c in couples]}")
#             print(f"Trainers: {[t.name for t in trainers]}")
#             print(f"{'='*60}")

#             # Step 1: Create trainer windows
#             try:
#                 tw = make_trainers_windows(trainers)
#                 print(f"✓ Trainer windows created (before filtering)")
#                 for trainer, windows in tw.items():
#                     available_slots = sum(1 for _, is_avail in windows if is_avail)
#                     print(f"  {trainer.name}: {available_slots} available slots")
#             except Exception as e:
#                 print(f"✗ Error creating trainer windows: {e}")
#                 continue

#             # Step 2: Sort couples
#             if sort_couple_by == 'group':
#                 sc = sort_couples_by_group(couples, Group.objects.all())
#             else:
#                 sc = list(couples)
#             print(f"✓ Couples sorted: {[c.name for c in sc]}")

#             # Step 3: Read dancers availability and convert to couple availability
#             couples_availability = {}
#             day_times_from_excel = None

#             for group in Group.objects.all():
#                 try:
#                     day_times, day_dancers_avail = read_dancers_availability(
#                         group.name,
#                         uploaded_file.file.path
#                     )

#                     if day.name not in day_times:
#                         print(f"  Group {group.name}: '{day.name}' not found in Excel")
#                         continue

#                     print(f"✓ Group {group.name}: Found {day.name}")
                    
#                     all_times = day_times[day.name]
#                     print(f"  All times in Excel: {all_times}")
#                     print(f"  Filtering to day hours: {day.start_time} - {day.end_time}")
                    
#                     # FILTER TIMES TO DAY HOURS
#                     filtered_times, filtered_avail_lists = filter_times_by_day_hours(
#                         all_times,
#                         day_dancers_avail[day.name],
#                         day.start_time,
#                         day.end_time
#                     )
                    
#                     print(f"  Filtered times: {filtered_times}")
                    
#                     if not filtered_times:
#                         print(f"  ✗ No times available within day hours for {group.name}")
#                         continue
                    
#                     day_times_from_excel = filtered_times
                    
#                     # Filter dancers' availability arrays
#                     filtered_dancers_avail = {}
#                     for dancer_name, avail_list in day_dancers_avail[day.name].items():
#                         # Create a mapping of times to availability
#                         time_to_avail = {}
#                         for idx, t in enumerate(all_times):
#                             if idx < len(avail_list):
#                                 time_to_avail[t] = avail_list[idx]
                        
#                         # Extract availability for filtered times only
#                         filtered_dancers_avail[dancer_name] = [
#                             time_to_avail.get(t, False) for t in filtered_times
#                         ]
                    
#                     print(f"  Dancers: {list(filtered_dancers_avail.keys())}")

#                     # Convert dancers availability to couples availability
#                     couple_avail = convert_dancers_to_couples_availability(couples, filtered_dancers_avail)

#                     if couple_avail:
#                         couples_availability.update(couple_avail)
#                         print(f"  ✓ Couple availability created: {list(couple_avail.keys())}")

#                 except Exception as e:
#                     print(f"  ✗ Error reading group {group.name}: {e}")
#                     import traceback
#                     traceback.print_exc()
#                     continue

#             if not couples_availability or not day_times_from_excel:
#                 print(f"✗ No couple availability data found for {day.name}")
#                 continue

#             # Step 4: Convert to availability with times
#             try:
#                 cawt_result = cawt(day_times_from_excel, couples_availability)
#                 print(f"✓ Availability with times created")
#                 print(f"  Available times for couples:")
#                 for couple_name, avail_with_times in cawt_result.items():
#                     available_times = [str(t) for t, avail in avail_with_times if avail]
#                     print(f"    {couple_name}: {available_times if available_times else 'NONE'}")
#             except Exception as e:
#                 print(f"✗ Error converting to times: {e}")
#                 import traceback
#                 traceback.print_exc()
#                 continue

#             # Step 5: Create schedule via backtracking
#             try:
#                 print(f"\nAttempting to create schedule...")
#                 schedule = create_the_schedule(
#                     trainers_windows=tw,
#                     sorted_couples=sc,
#                     couples_avail_with_times=cawt_result
#                 )

#                 if schedule:
#                     all_schedules[day.name] = schedule
#                     print(f"✓ Schedule created for {day.name}")
#                     print(f"  Scheduled couples:")
#                     for couple, (trainer, time_str) in schedule.items():
#                         print(f"    {couple.name}: {time_str} with {trainer.name}")
#                 else:
#                     print(f"✗ Backtracking failed for {day.name}")

#             except Exception as e:
#                 print(f"✗ Error during backtracking: {e}")
#                 import traceback
#                 traceback.print_exc()
#                 continue

#         if not all_schedules:
#             return JsonResponse({
#                 'status': 'error',
#                 'message': 'Could not create schedule for any configured day. Check the console output for details.'
#             }, status=400)

#         # Step 6: Format output
#         formatted_schedule = {}
#         for day_name, day_schedule in all_schedules.items():
#             formatted_schedule[day_name] = []
#             for couple, (trainer, time_str) in day_schedule.items():
#                 formatted_schedule[day_name].append({
#                     'couple': couple.name,
#                     'trainer': trainer.name,
#                     'time': time_str,
#                     'duration': couple.min_duration,
#                     'class_stt': couple.dance_class_stt,
#                     'class_lat': couple.dance_class_lat,
#                 })

#         return JsonResponse({
#             'status': 'success',
#             'message': 'Schedule created successfully',
#             'schedule': formatted_schedule
#         })

#     return JsonResponse({
#         'status': 'error',
#         'message': 'Invalid request method'
#     }, status=400)
            
